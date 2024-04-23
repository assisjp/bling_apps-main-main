import os
import time
import json
import secrets
import openpyxl
import requests
import base64
import sqlalchemy
import shutil
import csv
import re
from datetime import datetime, timedelta
from functools import wraps
from dotenv import load_dotenv
from flask import Flask, request, redirect, jsonify, render_template, session
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.exc import IntegrityError
from sqlalchemy_utils import StringEncryptedType
from sqlalchemy_utils.types.encrypted.encrypted_type import AesEngine
from werkzeug.security import generate_password_hash, check_password_hash
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
from markdown2 import markdown
from flask_session import Session
from flask import send_file


# Configurações iniciais
app = Flask(__name__)
load_dotenv()

# Configurações do Flask e SQLite
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///banco.db'
app.config['SESSION_TYPE'] = 'filesystem'
db = SQLAlchemy(app)
Session(app)

# Constantes
SECRET_KEY = os.getenv('SECRET_KEY')
REDIRECT_URI = 'http://localhost:5000/callback'
AUTH_URL = 'https://bling.com.br/Api/v3/oauth/authorize'
TOKEN_URL = 'https://bling.com.br/Api/v3/oauth/token'
STATE = secrets.token_hex(16)
BASE_URL = 'https://bling.com.br/Api/v3'
CONCURRENCY_LIMIT = 2
WAIT_INTERVAL = 0.5  # 500ms
NUVEMSHOP_URL = 'https://api.tiendanube.com/v1'
NUVEMSHOP_TOKEN_URL = 'https://www.nuvemshop.com.br/apps/authorize/token'
TEMP_CSV_DIR = 'temp_csv_files'
os.makedirs(TEMP_CSV_DIR, exist_ok=True)

errors = []

# Modelos
class User(db.Model):
    id = db.Column(db.String(32), primary_key=True, default=lambda: secrets.token_hex(16))
    username = db.Column(db.String(50), unique=True, nullable=False)
    password = db.Column(StringEncryptedType(db.String, SECRET_KEY, AesEngine, 'pkcs5'), nullable=False)
    tokenbling = db.relationship('Tokenbling', backref='user', lazy=True)
    tokennuvem = db.relationship('Tokennuvem', backref='user', lazy=True)

    def __init__(self, username, password):
        self.username = username
        self.password = generate_password_hash(password)

class Tokenbling(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.String(32), db.ForeignKey('user.id'), nullable=False)
    client_id = db.Column(StringEncryptedType(db.String, SECRET_KEY, AesEngine, 'pkcs5'), unique=True, nullable=True)
    client_secret = db.Column(StringEncryptedType(db.String, SECRET_KEY, AesEngine, 'pkcs5'), unique=True, nullable=True)
    access_token = db.Column(StringEncryptedType(db.String, SECRET_KEY, AesEngine, 'pkcs5'), unique=True, nullable=True, default=None)
    refresh_token = db.Column(StringEncryptedType(db.String, SECRET_KEY, AesEngine, 'pkcs5'), nullable=True, default=None)
    token_creation_time = db.Column(db.String(120), nullable=True)
    token_expiration_time = db.Column(db.String(120), nullable=True)

class Tokennuvem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.String(32), db.ForeignKey('user.id'), nullable=False)
    client_id_nuvemshop = db.Column(StringEncryptedType(db.String, SECRET_KEY, AesEngine, 'pkcs5'), unique=True, nullable=True, default=None)
    client_secret_nuvemshop = db.Column(StringEncryptedType(db.String, SECRET_KEY, AesEngine, 'pkcs5'), unique=True, nullable=True, default=None)
    access_token_nuvemshop = db.Column(StringEncryptedType(db.String, SECRET_KEY, AesEngine, 'pkcs5'), unique=True, nullable=True, default=None)
    code_nuvemshop = db.Column(StringEncryptedType(db.String, SECRET_KEY, AesEngine, 'pkcs5'), unique=True, nullable=True, default=None)
    refresh_token_nuvemshop = db.Column(StringEncryptedType(db.String, SECRET_KEY, AesEngine, 'pkcs5'), nullable=True, default=None)
    store_id_nuvemshop = db.Column(db.String(120), nullable=True)
    token_creation_time_nuvemshop = db.Column(db.String(120), nullable=True)
    token_expiration_time_nuvemshop = db.Column(db.String(120), nullable=True)

class Productsnuvemshop(db.Model):
    variant_id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    price = db.Column(db.Float, nullable=True)
    stock_management = db.Column(db.Boolean, nullable=True)
    created_at = db.Column(db.DateTime, nullable=False)
    user_id = db.Column(db.String(32), db.ForeignKey('user.id'), nullable=False)
    sku = db.Column(db.String(50), nullable=True)
    product_id_pai = db.Column(db.Integer, nullable=True)

class Descricaoproduto(db.Model):
    sku = db.Column(db.String(50), primary_key=True)
    descricao = db.Column(db.String(1000), nullable=False)
    site = db.Column(db.String(1000), nullable=True)
    marketplace = db.Column(db.String(1000), nullable=True)

class Productsbling(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    data = db.Column(db.String, nullable=False)  # Armazena os dados do produto como string JSON
    creation_time = db.Column(db.DateTime, default=datetime.utcnow)
    user_id = db.Column(db.String(32), db.ForeignKey('user.id'), nullable=False)

# Criação das tabelas    
@app.before_request
def create_tables():
    db.create_all()

# Funções auxiliares
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect('/login')
        return f(*args, **kwargs)
    return decorated_function

def replace_with_config_value(template, row_data):
    if isinstance(template, str):
        # Encontrar todas as chaves de substituição no template_str
        keys = re.findall(r"\{\{(.*?)\}\}", template)

        # Substituir cada chave pelo valor correspondente em row_data
        for key in keys:
            value = row_data.get(key, "")
            template = template.replace("{{" + key + "}}", str(value))

        return template
    elif isinstance(template, dict):
        # Se o template é um dicionário, processar cada campo recursivamente
        return {k: replace_with_config_value(v, row_data) for k, v in template.items()}
    elif isinstance(template, list):
        # Se o template é uma lista, processar cada item recursivamente
        return [replace_with_config_value(item, row_data) for item in template]
    else:
        # Se o template não é uma string, dicionário ou lista, retorná-lo como está
        return template

def refresh_token():
    users = User.query.all()
    for user in users:
        token_entry = user.tokenbling[0]
        if not token_entry or not token_entry.refresh_token:
            print(f"Refresh token not found for user {user.username}.")
            continue
        credentials = base64.b64encode(f"{token_entry.client_id}:{token_entry.client_secret}".encode()).decode()
        headers = {
            'Authorization': f"Basic {credentials}",
            'Content-Type': 'application/x-www-form-urlencoded',
            'Accept': '1.0'
        }
        refresh_data = {
            'grant_type': 'refresh_token',
            'refresh_token': token_entry.refresh_token
        }
        response = requests.post(TOKEN_URL, data=refresh_data, headers=headers)
        tokens = response.json()
        if 'access_token' in tokens:
            token_creation_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            token_expiration_time = (datetime.now() + timedelta(seconds=tokens['expires_in'])).strftime('%Y-%m-%d %H:%M:%S')
            token_entry.access_token = tokens['access_token']
            token_entry.refresh_token = tokens.get('refresh_token', None)
            token_entry.token_creation_time = token_creation_time
            token_entry.token_expiration_time = token_expiration_time
            db.session.commit()
        else:
            print(f"Unexpected response from API for user {user.username}: {tokens}")

def clear_old_products():
    with app.app_context():
        current_time = datetime.utcnow()
        threshold_time = current_time - timedelta(hours=24)
        print(f"Current time: {current_time}")
        print(f"Threshold time: {threshold_time}")
        
        old_products = Productsbling.query.filter(Productsbling.creation_time < threshold_time).all()
        for product in old_products:
            db.session.delete(product)
        db.session.commit()
        print(f"Deleted {len(old_products)} old products.")

def backup_database():
    print("Realizando backup do banco de dados...")
    try:
        backup_path = f"backup_path/backup_{datetime.now().strftime('%Y%m%d%H%M%S')}.db"
        shutil.copyfile('instance/banco.db', backup_path)
        print(f"Backup realizado com sucesso: {backup_path}")
    except Exception as e:
        print(f"Erro ao realizar backup: {str(e)}")


def json_to_csv(json_data, user_id, table_id):
    csv_filename = f'{TEMP_CSV_DIR}/products_table_{table_id}_user_{user_id}.csv'
    with open(csv_filename, mode='w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        # Escrevendo cabeçalhos do CSV
        writer.writerow(['ID', 'Nome', 'Código', 'Preço', 'Tipo', 'Situação', 'Formato', 'Descrição Curta'])

        # Escrevendo linhas de dados
        for item in json_data['data']:
            writer.writerow([
                item.get('id', ''),
                item.get('nome', ''),
                item.get('codigo', ''),
                item.get('preco', ''),
                item.get('tipo', ''),
                item.get('situacao', ''),
                item.get('formato', ''),
                item.get('descricaoCurta', '')
            ])
    return csv_filename


def clean_old_csv_files():
    now = datetime.now()
    for filename in os.listdir(TEMP_CSV_DIR):
        file_path = os.path.join(TEMP_CSV_DIR, filename)
        if os.path.isfile(file_path):
            creation_time = datetime.fromtimestamp(os.path.getctime(file_path))
            if (now - creation_time).total_seconds() > 300:  # 300 segundos = 5 minutos
                try:
                    os.remove(file_path)
                    print(f"Removed old CSV file: {filename}")
                except OSError as e:
                    print(f"Error removing file {filename}: {e}")


# Rotas
@app.route('/preencher-descricao', methods=['GET', 'POST'])
def preencher_descricao():
    if request.method == 'POST':
        sku = request.form.get('sku')
        descricao = request.form.get('descricao')
        site = request.form.get('site')
        marketplace = request.form.get('marketplace')
        if sku and descricao and site and marketplace:
            descricaoproduto = Descricaoproduto(sku=sku, descricao=descricao, site=site, marketplace=marketplace)
            db.session.add(descricaoproduto)
            db.session.commit()
            return redirect('/preencher-descricao')
        else:
            return render_template('preencher_descricao.html', error='Todos os campos são obrigatórios')
    else:
        return render_template('preencher_descricao.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        if username and password:
            user = User.query.filter_by(username=username).first()
            if user and check_password_hash(user.password, password):
                session['user_id'] = user.id
                return redirect('/dashboard')
            else:
                return render_template('login.html', error='Usuário ou senha incorretos')
        else:
            return render_template('login.html', error='Usuário e senha são obrigatórios')
    else:
        return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')

@app.route('/create_user', methods=['GET', 'POST'])
def create_user():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        client_id = request.form.get('client_id')
        client_secret = request.form.get('client_secret')
        if username and password:
            user = User(username=username, password=password)
            db.session.add(user)
            try:
                db.session.commit()
            except IntegrityError:
                return render_template('createuser.html', error='Usuário já existe')
            Tokennuvem(user_id=user.id)
            token = Tokenbling(user_id=user.id, client_id=client_id or None, client_secret=client_secret or None)
            db.session.add(token)
            db.session.commit()
            return redirect('/login')
        else:
            return render_template('createuser.html', error='Usuário e senha são obrigatórios')
    else:
        return render_template('createuser.html')

@app.route('/')
@login_required
def index():
    return render_template('dashboard.html')

@app.route('/auth', methods=['GET'])
@login_required
def authenticate():
    user_id = session['user_id']
    user = User.query.get(user_id)
    client_id = user.tokenbling[0].client_id
    auth_url = f"{AUTH_URL}?response_type=code&client_id={client_id}&state={STATE}"
    return redirect(auth_url)

@app.route('/callback', methods=['GET'])
@login_required
def callback():
    code = request.args.get('code')
    user_id = session['user_id']
    user = User.query.get(user_id)
    token_entry = user.tokenbling[0]
    credentials = base64.b64encode(f"{token_entry.client_id}:{token_entry.client_secret}".encode()).decode()
    headers = {
        'Authorization': f"Basic {credentials}",
        'Content-Type': 'application/x-www-form-urlencoded',
        'Accept': '1.0'
    }
    token_data = {
        'grant_type': 'authorization_code',
        'code': code
    }
    response = requests.post(TOKEN_URL, data=token_data, headers=headers)
    response_data = response.json()
    if 'access_token' in response_data:
        token_creation_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        token_expiration_time = (datetime.now() + timedelta(seconds=response_data['expires_in'])).strftime('%Y-%m-%d %H:%M:%S')
        token_entry.access_token = response_data['access_token']
        token_entry.refresh_token = response_data.get('refresh_token', None)
        token_entry.token_creation_time = token_creation_time
        token_entry.token_expiration_time = token_expiration_time
        db.session.commit()
        return redirect('/dashboard')
    else:
        return jsonify({"error": "Unexpected response from the server"}), 500

@app.route('/dashboard', methods=['GET'])
@login_required
def dashboard():
    user_id = session['user_id']
    user = User.query.get(user_id)
    token_entry = user.tokenbling[0]
    if token_entry:
        token_status = token_entry.access_token is not None and datetime.now() < datetime.strptime(token_entry.token_expiration_time, '%Y-%m-%d %H:%M:%S')
        return render_template('dashboard.html', token_status=token_status, last_updated=datetime.now().strftime('%Y-%m-%d %H:%M:%S'), created_at=token_entry.token_creation_time, expires_at=token_entry.token_expiration_time)
    else:
        return render_template('dashboard.html', token_status=False, last_updated=None, created_at=None, expires_at=None)

@app.route('/validate-token', methods=['GET'])
@login_required
def validate_token():
    try:
        refresh_token()
        return redirect('/dashboard')
    except:
        return redirect('/dashboard')

@app.route('/products', methods=['GET'])
@login_required
def get_products():
    print("Getting products from API...")
    user_id = session.get('user_id')
    token = db.session.query(Tokenbling).filter_by(user_id=user_id).first()
    token_entry = Tokenbling.query.filter_by(user_id=user_id).first()
    if not user_id:
        return jsonify({"error": "User not logged in"}), 401
    access_token = token_entry.access_token if token_entry else None
    if not access_token:
        return jsonify({"error": "Access token not found"}), 401

    headers = {
        'Authorization': f"Bearer {token.access_token}"
    }

    params = {}
    for key in ['integrate', 'limit', 'offset', 'image', 'loja', 'estoque', 'tipo']:
        value = request.args.get(key)
        if value:
            params[key] = value

    response = requests.get(f"{BASE_URL}/produtos", headers=headers, params=params)
    data = response.json()

    if response.status_code != 200:
        return jsonify({"error": f"API returned {response.status_code}: {response.text}"}), 500

    # Salvar os produtos no banco de dados
    product_entry = Productsbling(data=json.dumps(data), user_id=user_id) 
    db.session.add(product_entry)
    db.session.commit()

    # Gerar um link local para acessar os produtos
    local_link = f"/local-products/{product_entry.id}"

    return jsonify({"message": "Products fetched successfully", "local_link": local_link})

@app.route('/list-products-tables')
@login_required
def list_products_tables():
    user_id = session.get('user_id')
    try:
        product_entries = Productsbling.query.filter_by(user_id=user_id).all()
        if not product_entries:
            return render_template('list_tables.html', message="Nenhuma tabela de produtos encontrada.")

        # Criando uma lista de IDs de tabela para o usuário escolher
        table_ids = [entry.id for entry in product_entries]

        return render_template('list_tables.html', tables=table_ids)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/download-products-csv/<int:table_id>')
@login_required
def download_products_csv(table_id):
    user_id = session.get('user_id')
    try:
        product_entry = Productsbling.query.filter_by(user_id=user_id, id=table_id).first()
        if not product_entry:
            return jsonify({"error": "Product table not found"}), 404

        product_data = json.loads(product_entry.data)
        csv_filename = f'products_table_{table_id}_user_{user_id}.csv'
        csv_filepath = os.path.join(TEMP_CSV_DIR, csv_filename)

        with open(csv_filepath, 'w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            # Escrevendo cabeçalhos do CSV
            writer.writerow(['ID', 'Nome', 'Código', 'Preço', 'Tipo', 'Situação', 'Formato', 'Descrição Curta'])
            # Escrevendo linhas de dados
            for item in product_data['data']:
                writer.writerow([
                    item.get('id', ''),
                    item.get('nome', ''),
                    item.get('codigo', ''),
                    item.get('preco', ''),
                    item.get('tipo', ''),
                    item.get('situacao', ''),
                    item.get('formato', ''),
                    item.get('descricaoCurta', '')
                ])

        return send_file(csv_filepath, as_attachment=True, mimetype='text/csv', download_name=csv_filename)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/local-products/<int:product_id>', methods=['GET'])
@login_required
def get_local_products(product_id):
    product_entry = Productsbling.query.get(product_id)
    if not product_entry:
        return jsonify({"error": "Product not found"}), 404
    data = json.loads(product_entry.data)
    return jsonify({"data": data})

@app.route('/upload-excel', methods=['POST'])
@login_required
def upload_excel():
    user_id = session.get('user_id')
    token_entry = Tokenbling.query.filter_by(user_id=user_id).first()
    if not token_entry:
        return jsonify({"error": "Access token not found"}), 401

    if 'file' not in request.files:
        return jsonify({"error": "No file part"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400

    # Carregar a configuração do cliente
    client_config_path = os.path.join("client_configs", f"{user_id}.json")
    if not os.path.exists(client_config_path):
        return jsonify({"error": "Client configuration not found"}), 400

    with open(client_config_path, 'r') as config_file:
        client_config = json.load(config_file)

    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active

    headers = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]

    errors = []
    products = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        product = {}
        row_data = dict(zip(headers, row))

        # Processar cada campo de acordo com o client_config
        for key, template in client_config.items():
            if isinstance(template, dict):  # Verificar se o template é um objeto JSON
                product[key] = {}
                for sub_key, sub_template in template.items():
                    if "{{" in str(sub_template):
                        product[key][sub_key] = replace_with_config_value(sub_template, row_data)
                    else:
                        product[key][sub_key] = sub_template
            else:
                if "{{" in str(template):
                    product[key] = replace_with_config_value(template, row_data)
                else:
                    product[key] = template

        # Agrupar as variações por produto pai
        codigo = product.get('codigo')
        if codigo not in products:
            products[codigo] = product
        else:
            products[codigo]['variacoes'].extend(product['variacoes'])

    headers_request = {
        'Authorization': f"Bearer {token_entry.access_token}",
        'Content-Type': 'application/json'
    }

    for product in products.values():
        print(f"Sending product: {product}")

        # Ajustar o envio do produto de acordo com o formato da API REST
        response = requests.post(
            f"{BASE_URL}/produtos",
            #"https://typedwebhook.tools/webhook/5cf4073a-54ea-45df-9ab1-c5cc2ffe7285",
            headers=headers_request,
            data=json.dumps(product)  # Enviar o JSON como string diretamente
        )

        print(f"Response: {response.text}, Status code: {response.status_code}")

        if response.status_code != 200:
            errors.append(product.get('codigo', 'Unknown'))

        time.sleep(WAIT_INTERVAL)

    if errors:
        return jsonify({"error": f"Failed to add products {', '.join(errors)}"}), 500

    return jsonify({"message": "Products added successfully"}), 200

@app.errorhandler(500)
def internal_error(error):
    return "Erro interno do servidor", 500

@app.errorhandler(404)
def not_found(error):
    return "Página não encontrada", 404

@app.route('/configure-bling', methods=['GET', 'POST'])
@login_required
def configure_bling():
    user_id = session['user_id']
    user = User.query.get(user_id)
    token_bling = user.tokenbling[0] if user.tokenbling else None

    if request.method == 'POST':
        client_id_bling = request.form.get('client_id_bling')
        client_secret_bling = request.form.get('client_secret_bling')

        if token_bling:
            token_bling.client_id = client_id_bling or token_bling.client_id
            token_bling.client_secret = client_secret_bling or token_bling.client_secret
        else:
            token_bling = Tokenbling(user_id=user.id, client_id=client_id_bling, client_secret=client_secret_bling)
            db.session.add(token_bling)

        db.session.commit()
        return redirect('/dashboard')

    return render_template('configure_user.html', token_bling=token_bling)


@app.route('/configure-nuvemshop', methods=['GET', 'POST'])
@login_required
def configure_nuvemshop():
    user_id = session['user_id']
    user = User.query.get(user_id)
    token_entry = user.tokennuvem[0] if user.tokennuvem else None

    if user.tokennuvem:
        token_entry = user.tokennuvem[0]

    if request.method == 'POST':
        client_id_nuvemshop = request.form.get('client_id_nuvemshop')
        client_secret_nuvemshop = request.form.get('client_secret_nuvemshop')
        code_nuvemshop = request.form.get('code_nuvemshop')
        
        if token_entry:
            token_entry.client_id_nuvemshop = client_id_nuvemshop
            token_entry.client_secret_nuvemshop = client_secret_nuvemshop
            token_entry.code_nuvemshop = code_nuvemshop
        else:
            token_entry = Tokennuvem(user_id=user.id, client_id_nuvemshop=client_id_nuvemshop, client_secret_nuvemshop=client_secret_nuvemshop, code_nuvemshop=code_nuvemshop)
            db.session.add(token_entry)

        db.session.commit()

        return redirect('/dashboard-nuvemshop')

    return render_template('configure_user.html', client_id_nuvemshop=token_entry.client_id_nuvemshop if token_entry else None, client_secret_nuvemshop=token_entry.client_secret_nuvemshop if token_entry else None, code_nuvemshop=token_entry.code_nuvemshop if token_entry else None)

@app.route('/callback-nuvemshop', methods=['GET'])
@login_required
def callback_nuvemshop():
    user_id = session['user_id']
    user = User.query.get(user_id)
    token_entry = user.tokennuvem[0]

    # Preparar os dados para a requisição de token
    token_data = {
        'client_id': token_entry.client_id_nuvemshop,
        'client_secret': token_entry.client_secret_nuvemshop,
        'grant_type': 'authorization_code',
        'code': token_entry.code_nuvemshop
    }

    # Fazer a requisição para obter o token de acesso
    headers = {
        'Content-Type': 'application/json',
        'User-Agent': 'MyApp test@tiendanube.com'
    }
    response = requests.post(NUVEMSHOP_TOKEN_URL, headers=headers, json=token_data)

    # Verificar se a resposta foi bem-sucedida
    if response.status_code == 200:
        response_data = response.json()
        if 'access_token' in response_data:
            token_entry.access_token_nuvemshop = response_data['access_token']
            token_entry.store_id_nuvemshop = response_data['user_id']
            token_entry.token_creation_time_nuvemshop = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            db.session.commit()
            return redirect('/dashboard-nuvemshop')
        else:
            # Retornar a resposta como HTML
            return response.content, 500, {'Content-Type': 'text/html'}
    else:
        # Retornar a resposta como HTML
        return response.content, 500, {'Content-Type': 'text/html'}

    
@app.route('/dashboard-nuvemshop')
@login_required
def dashboard_nuvemshop():
    user_id = session['user_id']
    user = User.query.get(user_id)

    # Verifique se há algum token associado ao usuário
    token_entry = user.tokennuvem[0] if user.tokennuvem else None

    if token_entry:
        token_status_nuvemshop = token_entry.access_token_nuvemshop is not None
        return render_template('dashboard_nuvemshop.html', token_status=token_status_nuvemshop, last_updated=datetime.now().strftime('%Y-%m-%d %H:%M:%S'), created_at=token_entry.token_creation_time_nuvemshop)
    else:
        return render_template('dashboard_nuvemshop.html', token_status=False, last_updated=None, created_at=None)

@app.route('/productsnuvemshop', methods=['GET'])
@login_required
def fetch_and_save_products():
    """
    Fetches products from the Nuvemshop API and saves them to the database.

    Returns:
        JSON response: A JSON response indicating the success or failure of the operation.
    """
    print("Getting products from API...")
    user_id = session.get('user_id')
    token_entry = Tokennuvem.query.filter_by(user_id=user_id).first()
    store_id = token_entry.store_id_nuvemshop if token_entry else None
    if not user_id:
        return jsonify({"error": "User not logged in"}), 401
    access_token_nuvemshop = token_entry.access_token_nuvemshop if token_entry else None
    if not access_token_nuvemshop:
        return jsonify({"error": "Access token not found"}), 401

    headers = {
        'Authentication': f"bearer {access_token_nuvemshop}",
        'User-Agent': 'MyApp test@tiendanube.com'
    }
    
    page = 1
    per_page = 200
    all_products_fetched = False

    while not all_products_fetched:
        params = {
            'per_page': per_page,
            'page': page
        }

        response = requests.get(f"{NUVEMSHOP_URL}/{store_id}/products", headers=headers, params=params)
        if response.status_code == 404:
            break  # Para o loop se não houver mais páginas
        elif response.status_code != 200:
            app.logger.error(f"Failed to fetch products from Nuvemshop: {response.text}")
            return jsonify({"error": "Failed to fetch products from Nuvemshop", "details": response.text}), response.status_code

        products_data = response.json()
        if not products_data:  # Se a resposta estiver vazia, para o loop
            break

        for product in products_data:
            product_name = product['name']['pt']
            for variant in product['variants']:
                variant_id = variant['id']
                existing_variant = Productsnuvemshop.query.filter_by(variant_id=variant_id).first()
                
                if existing_variant:
                    # Atualiza o registro existente
                    existing_variant.name = product_name
                    existing_variant.sku = variant.get('sku')
                    existing_variant.price = variant['price'] if variant.get('price') not in [None, 'None'] else None
                    existing_variant.stock_management = variant['stock_management']
                    existing_variant.created_at = datetime.strptime(variant['created_at'], "%Y-%m-%dT%H:%M:%S%z").replace(tzinfo=None)
                    existing_variant.user_id = user_id
                    existing_variant.product_id_pai = product['id']
                else:
                    # Cria um novo registro
                    new_variant = Productsnuvemshop(
                        variant_id=variant_id,
                        name=product_name,
                        sku=variant.get('sku'),
                        price=variant['price'] if variant.get('price') not in [None, 'None'] else None,
                        stock_management=variant['stock_management'],
                        created_at=datetime.strptime(variant['created_at'], "%Y-%m-%dT%H:%M:%S%z").replace(tzinfo=None),
                        user_id=user_id,
                        product_id_pai=product['id']
                    )
                    db.session.add(new_variant)
            pass

        page += 1  # Incrementa o número da página

    try:
        db.session.commit()
        return jsonify({"message": "Products fetched and saved successfully"}), 200
    except sqlalchemy.exc.IntegrityError as e:
        db.session.rollback()
        app.logger.error(f"IntegrityError: {str(e)}")
        return jsonify({"error": "Failed to commit to the database", "details": str(e)}), 500
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Unknown error: {str(e)}")
        return jsonify({"error": "Unknown error", "details": str(e)}), 500


@app.route('/update-metafields', methods=['POST'])
@login_required
def update_metafields():
    user_id = session.get('user_id')
    token_entry = Tokennuvem.query.filter_by(user_id=user_id).first()
    access_token_nuvemshop = token_entry.access_token_nuvemshop if token_entry else None
    store_id = token_entry.store_id_nuvemshop if token_entry else None

    if not access_token_nuvemshop:
        return jsonify({"error": "Access token not found"}), 401

    all_products = Productsnuvemshop.query.all()
    processed_products = []

    for product in all_products:
        description_entry = Descricaoproduto.query.filter_by(sku=product.sku).first()
        if description_entry:
            # Converter para HTML usando Markdown
            description_html = markdown(description_entry.descricao)
            # description_html = description_html.replace('<p>', '', 1)  # Remover a primeira tag <p>
            description_html = description_html.replace('\n', '')  # Remover quebras de linha

            metafield_data = {
                'namespace': 'descriptions',
                'key': 'variant_id_' + str(product.variant_id),
                'value': description_html,
                'owner_id': product.product_id_pai,
                'owner_resource': 'Product'
            }

            headers = {
                'Content-Type': 'application/json',
                'Authentication': f"bearer {access_token_nuvemshop}",
                'User-Agent': 'MyApp test@tiendanube.com'
            }

            #response = requests.post(f"{NUVEMSHOP_URL}/{store_id}/metafields", headers=headers, json=metafield_data)
            response = requests.post(f"https://webhook.site/7b7048a6-7da1-4c9e-ae8f-921bd35c3c31/{store_id}/metafields", headers=headers, json=metafield_data)

            if response.status_code != 201:
                print(f"Erro ao atualizar metafield para o produto {product.name} com SKU {product.sku}: {response.content}")
            else:
                processed_products.append({"name": product.name, "sku": product.sku})

    if processed_products:
        return jsonify({"message": "Metafields atualizados com sucesso", "processed_products": processed_products}), 200
    else:
        return jsonify({"message": "Nenhum produto processado"}), 200


scheduler = BackgroundScheduler()
scheduler.add_job(refresh_token, 'interval', minutes=340)
scheduler.add_job(clear_old_products, 'interval', minutes=10)
scheduler.add_job(backup_database, CronTrigger(hour=14, minute=55))
scheduler.add_job(clean_old_csv_files, 'interval', minutes=10)
scheduler.start()

if __name__ == '__main__':
    #db.create_all()
    app.run(host="0.0.0.0", port=5000, debug=False)